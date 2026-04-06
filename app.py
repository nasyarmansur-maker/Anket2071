from flask import Flask, render_template, request, redirect, url_for, session, send_file, jsonify
import sqlite3, json, os, io, base64, smtplib, qrcode
from datetime import datetime
from functools import wraps
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
app.secret_key = "okul_anket_super_gizli_2024"
DB = "anket.db"

# ─── Veritabanı ───────────────────────────────────────────────────
def db():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    return con

def db_init():
    with db() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS ayarlar (
            anahtar TEXT PRIMARY KEY, deger TEXT
        );
        CREATE TABLE IF NOT EXISTS anketler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rol TEXT NOT NULL, baslik TEXT NOT NULL,
            icon TEXT NOT NULL DEFAULT '📝',
            aktif INTEGER DEFAULT 1, sira INTEGER DEFAULT 0,
            baslangic_tarihi TEXT DEFAULT NULL,
            bitis_tarihi TEXT DEFAULT NULL,
            aciklama TEXT DEFAULT '',
            gorunum TEXT DEFAULT 'varsayilan'
        );
        CREATE TABLE IF NOT EXISTS bolumler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            anket_id INTEGER NOT NULL, baslik TEXT NOT NULL, sira INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS sorular (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bolum_id INTEGER NOT NULL, metin TEXT NOT NULL,
            tip TEXT NOT NULL DEFAULT 'yildiz',
            secenekler TEXT DEFAULT '[]',
            zorunlu INTEGER DEFAULT 1, sira INTEGER DEFAULT 0,
            kosul_soru_id INTEGER DEFAULT NULL,
            kosul_deger TEXT DEFAULT NULL
        );
        CREATE TABLE IF NOT EXISTS yanitlar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            anket_id INTEGER NOT NULL,
            tarih TEXT NOT NULL, saat TEXT NOT NULL, veriler TEXT NOT NULL
        );
        """)
        defaults = {
            "okul_adi": "Örnek İlköğretim Okulu",
            "okul_sehir": "Erzurum",
            "admin_sifre": "okul2024",
            "tema": "mavi",
            "amblem": "",
            "anasayfa_gorunum": "kartlar",
            "hosgeldin_metin": "Görüşleriniz okulumuzun gelişimine katkı sağlar.",
            "alt_yazi": "Anketler anonim olarak kaydedilir.",
            "smtp_host": "",
            "smtp_port": "587",
            "smtp_user": "",
            "smtp_pass": "",
            "bildirim_email": "",
            "bildirim_aktif": "0",
            "anket_varsayilan_gorunum": "adim",
            "anket_gorunum_secim_goster": "1"
        }
        for k, v in defaults.items():
            c.execute("INSERT OR IGNORE INTO ayarlar VALUES (?,?)", (k, v))
        if not c.execute("SELECT 1 FROM anketler").fetchone():
            _varsayilan_anketler(c)
        c.commit()

def _varsayilan_anketler(c):
    veri = [
        ("ogrenci","Öğrenci Değerlendirme Anketi","🎒",0,[
            ("Ders ve Öğretim",[
                ("Derslerin işlenişini nasıl değerlendirirsiniz?","yildiz","[]",1,None,None),
                ("Öğretmenler sorularınıza yeterince ilgi gösteriyor mu?","secim",'["Her zaman","Çoğunlukla","Bazen","Nadiren"]',1,None,None),
            ]),
            ("Okul Ortamı",[
                ("Okul temizliği ve düzenini nasıl buluyorsunuz?","yildiz","[]",1,None,None),
                ("Okul ortamında kendinizi güvende hissediyor musunuz?","secim",'["Evet, her zaman","Genellikle evet","Bazen hayır","Hayır"]',1,None,None),
            ]),
            ("Genel Değerlendirme",[
                ("Okulunuzdan genel memnuniyetiniz nedir?","yildiz","[]",1,None,None),
                ("Okulun daha iyi olması için öneriniz var mı?","metin","[]",0,None,None),
            ]),
        ]),
        ("veli","Veli Değerlendirme Anketi","👨‍👩‍👧",1,[
            ("İletişim ve Bilgilendirme",[
                ("Okul ile iletişim kalitesini nasıl değerlendirirsiniz?","yildiz","[]",1,None,None),
                ("Çocuğunuzun gelişimi hakkında yeterince bilgilendiriliyor musunuz?","secim",'["Evet, düzenli olarak","Kısmen","Hayır, yeterince değil"]',1,None,None),
            ]),
            ("Eğitim Kalitesi",[
                ("Okulun eğitim kalitesini nasıl değerlendirirsiniz?","yildiz","[]",1,None,None),
                ("Öğretmenler velilere karşı destekleyici mi?","secim",'["Evet","Çoğunlukla","Bazen","Hayır"]',1,None,None),
            ]),
            ("Genel Görüş",[
                ("Okuldan genel memnuniyetiniz?","yildiz","[]",1,None,None),
                ("Paylaşmak istediğiniz görüş veya öneri:","metin","[]",0,None,None),
            ]),
        ]),
        ("ogretmen","Öğretmen Değerlendirme Anketi","📚",2,[
            ("Yönetim ve Destek",[
                ("Okul yönetiminin öğretmenlere desteğini nasıl değerlendirirsiniz?","yildiz","[]",1,None,None),
                ("Okul kararlarında görüşleriniz dikkate alınıyor mu?","secim",'["Her zaman","Çoğunlukla","Bazen","Nadiren"]',1,None,None),
            ]),
            ("Çalışma Ortamı",[
                ("Fiziksel çalışma ortamını nasıl değerlendirirsiniz?","yildiz","[]",1,None,None),
                ("Ders materyalleri ve kaynaklar yeterli mi?","secim",'["Evet, çok yeterli","Kısmen yeterli","Yetersiz"]',1,None,None),
            ]),
            ("Genel Değerlendirme",[
                ("İş memnuniyetiniz genel olarak nasıl?","yildiz","[]",1,None,None),
                ("Okulu geliştirmek için önerileriniz:","metin","[]",0,None,None),
            ]),
        ]),
    ]
    for rol,baslik,icon,sira,bolumler in veri:
        aid=c.execute("INSERT INTO anketler (rol,baslik,icon,sira) VALUES (?,?,?,?)",(rol,baslik,icon,sira)).lastrowid
        for bs,(bb,sorular) in enumerate(bolumler):
            bid=c.execute("INSERT INTO bolumler (anket_id,baslik,sira) VALUES (?,?,?)",(aid,bb,bs)).lastrowid
            for ss,(sm,st,sse,sz,ksid,kdeger) in enumerate(sorular):
                c.execute("INSERT INTO sorular (bolum_id,metin,tip,secenekler,zorunlu,sira,kosul_soru_id,kosul_deger) VALUES (?,?,?,?,?,?,?,?)",
                          (bid,sm,st,sse,sz,ss,ksid,kdeger))

db_init()

# ─── Yardımcı ─────────────────────────────────────────────────────
def ayar(k, v=""):
    with db() as c:
        r=c.execute("SELECT deger FROM ayarlar WHERE anahtar=?",(k,)).fetchone()
    return r["deger"] if r else v

def ayar_set(k,v):
    with db() as c:
        c.execute("INSERT OR REPLACE INTO ayarlar VALUES (?,?)",(k,v)); c.commit()

def get_anket(aid):
    with db() as c:
        a=c.execute("SELECT * FROM anketler WHERE id=?",(aid,)).fetchone()
        if not a: return None
        bolumler=c.execute("SELECT * FROM bolumler WHERE anket_id=? ORDER BY sira",(aid,)).fetchall()
        result=dict(a); result["bolumler"]=[]
        for b in bolumler:
            bd=dict(b)
            sorular=c.execute("SELECT * FROM sorular WHERE bolum_id=? ORDER BY sira",(b["id"],)).fetchall()
            bd["sorular"]=[dict(s) for s in sorular]
            for s in bd["sorular"]:
                s["secenekler"]=json.loads(s["secenekler"] or "[]")
            result["bolumler"].append(bd)
        return result

def anket_aktif_mi(a):
    bugun = datetime.now().strftime("%Y-%m-%d")
    if a.get("baslangic_tarihi") and bugun < a["baslangic_tarihi"]:
        return False, "Anket henüz başlamadı."
    if a.get("bitis_tarihi") and bugun > a["bitis_tarihi"]:
        return False, "Anket süresi doldu."
    return True, ""

def gctx():
    return {"okul_adi":ayar("okul_adi"),"okul_sehir":ayar("okul_sehir"),
            "tema":ayar("tema","mavi"),"amblem":ayar("amblem")}

def giris_gerekli(f):
    @wraps(f)
    def k(*a,**kw):
        if not session.get("admin"): return redirect(url_for("admin_giris"))
        return f(*a,**kw)
    return k

def email_gonder(konu, icerik):
    try:
        if ayar("bildirim_aktif") != "1": return
        host=ayar("smtp_host"); port=int(ayar("smtp_port","587"))
        user=ayar("smtp_user"); pwd=ayar("smtp_pass")
        hedef=ayar("bildirim_email")
        if not all([host,user,pwd,hedef]): return
        msg=MIMEMultipart()
        msg["From"]=user; msg["To"]=hedef; msg["Subject"]=konu
        msg.attach(MIMEText(icerik,"html","utf-8"))
        with smtplib.SMTP(host,port) as s:
            s.starttls(); s.login(user,pwd); s.send_message(msg)
    except Exception as e:
        print(f"E-posta hatası: {e}")

# ─── Anasayfa ─────────────────────────────────────────────────────
@app.route("/")
def anasayfa():
    with db() as c:
        anketler=c.execute("SELECT * FROM anketler WHERE aktif=1 ORDER BY sira").fetchall()
    bugun=datetime.now().strftime("%Y-%m-%d")
    liste=[]
    for a in anketler:
        ad=dict(a)
        aktif,msg=anket_aktif_mi(ad)
        ad["simdiki_aktif"]=aktif; ad["mesaj"]=msg
        liste.append(ad)
    ctx=gctx(); ctx["anketler"]=liste
    ctx["hosgeldin"]=ayar("hosgeldin_metin"); ctx["alt_yazi"]=ayar("alt_yazi")
    ctx["gorunum"]=ayar("anasayfa_gorunum","kartlar")
    return render_template("anasayfa.html",**ctx)

# ─── Anket (adım adım) ────────────────────────────────────────────
@app.route("/anket/<int:anket_id>", methods=["GET","POST"])
def anket(anket_id):
    a=get_anket(anket_id)
    if not a or not a["aktif"]: return redirect(url_for("anasayfa"))
    aktif,msg=anket_aktif_mi(a)
    if not aktif:
        ctx=gctx(); ctx["mesaj"]=msg; ctx["anket"]=a
        return render_template("anket_kapali.html",**ctx)
    if request.method=="POST":
        v={k:request.form.getlist(k) if k.endswith("[]") else val
           for k,val in request.form.items()}
        # checkbox alanlarını düzgün topla
        veriler={}
        for key in request.form:
            vals=request.form.getlist(key)
            veriler[key]=",".join(vals) if len(vals)>1 else vals[0]
        with db() as c:
            c.execute("INSERT INTO yanitlar (anket_id,tarih,saat,veriler) VALUES (?,?,?,?)",
                      (anket_id,datetime.now().strftime("%d.%m.%Y"),
                       datetime.now().strftime("%H:%M"),
                       json.dumps(veriler,ensure_ascii=False)))
            c.commit()
        # E-posta bildirimi
        email_gonder(
            f"📋 Yeni Anket Yanıtı — {a['baslik']}",
            f"<h3>{a['icon']} {a['baslik']}</h3><p>Yeni bir yanıt alındı.</p>"
            f"<p><b>Tarih:</b> {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>"
        )
        return redirect(url_for("tesekkur",anket_id=anket_id))
    ctx=gctx(); ctx["anket"]=a
    ctx["sorular_json"]=json.dumps(
        {s["id"]:{"kosul_soru_id":s["kosul_soru_id"],"kosul_deger":s["kosul_deger"]}
         for b in a["bolumler"] for s in b["sorular"]},
        ensure_ascii=False)
    # Görünüm hesapla: URL param > kullanıcı tercihi > anket ayarı > sistem varsayılanı
    varsayilan = ayar("anket_varsayilan_gorunum","adim")
    anket_gorunum = a.get("gorunum","varsayilan")
    efektif = varsayilan if anket_gorunum=="varsayilan" else anket_gorunum
    url_gorunum = request.args.get("gorunum","")
    if url_gorunum in ["adim","tek"]: efektif = url_gorunum
    ctx["anket_gorunum"] = efektif
    ctx["gorunum_secim_goster"] = ayar("anket_gorunum_secim_goster","1") == "1"
    ctx["anket_varsayilan"] = varsayilan
    return render_template("anket.html",**ctx)

@app.route("/tesekkur/<int:anket_id>")
def tesekkur(anket_id):
    with db() as c:
        a=c.execute("SELECT * FROM anketler WHERE id=?",(anket_id,)).fetchone()
    ctx=gctx(); ctx["anket"]=dict(a) if a else {"baslik":"Anket","icon":"✅","id":0}
    return render_template("tesekkur.html",**ctx)

# ─── Admin giriş ─────────────────────────────────────────────────
@app.route("/admin",methods=["GET","POST"])
def admin_giris():
    hata=None
    if request.method=="POST":
        if request.form.get("sifre")==ayar("admin_sifre","okul2024"):
            session["admin"]=True; return redirect(url_for("admin_panel"))
        hata="Yanlış şifre!"
    return render_template("admin_giris.html",hata=hata,**gctx())

@app.route("/admin/cikis")
def admin_cikis():
    session.clear(); return redirect(url_for("anasayfa"))

# ─── Admin Panel (dashboard) ──────────────────────────────────────
@app.route("/admin/panel")
@giris_gerekli
def admin_panel():
    with db() as c:
        anketler=c.execute("SELECT * FROM anketler ORDER BY sira").fetchall()
        istat=[]
        toplam=0
        for a in anketler:
            sayi=c.execute("SELECT COUNT(*) as n FROM yanitlar WHERE anket_id=?",(a["id"],)).fetchone()["n"]
            toplam+=sayi
            yanitlar=c.execute("SELECT veriler FROM yanitlar WHERE anket_id=?",(a["id"],)).fetchall()
            # Yıldız ortalamaları
            yy={}
            for y in yanitlar:
                for k,val in json.loads(y["veriler"]).items():
                    if k.startswith("s_"):
                        try:
                            iv=int(val)
                            if 1<=iv<=5: yy.setdefault(k,[]).append(iv)
                        except: pass
            avg={k:round(sum(vs)/len(vs),1) for k,vs in yy.items()}
            # Son 7 gün istatistik
            son7=c.execute("""SELECT tarih, COUNT(*) as sayi FROM yanitlar
                WHERE anket_id=? GROUP BY tarih ORDER BY tarih DESC LIMIT 7""",(a["id"],)).fetchall()
            istat.append({"anket":dict(a),"sayi":sayi,"avg":avg,
                          "son7":[dict(r) for r in son7]})
        son=c.execute("""SELECT y.*,a.baslik as ab,a.icon as ai FROM yanitlar y
            JOIN anketler a ON y.anket_id=a.id ORDER BY y.id DESC LIMIT 10""").fetchall()
    ctx=gctx(); ctx.update({"istat":istat,"toplam":toplam,"son":[dict(r) for r in son]})
    return render_template("admin_panel.html",**ctx)

# ─── Admin Sonuçlar + filtreleme ──────────────────────────────────
@app.route("/admin/sonuclar/<int:anket_id>")
@giris_gerekli
def admin_sonuclar(anket_id):
    a=get_anket(anket_id)
    if not a: return redirect(url_for("admin_panel"))
    filtre_tarih=request.args.get("tarih","")
    filtre_arama=request.args.get("arama","").lower()
    with db() as c:
        q="SELECT * FROM yanitlar WHERE anket_id=?"
        params=[anket_id]
        if filtre_tarih:
            q+=" AND tarih=?"; params.append(filtre_tarih)
        q+=" ORDER BY id DESC"
        yanitlar=c.execute(q,params).fetchall()
        # Tüm tarihler (filtre için)
        tarihler=c.execute("SELECT DISTINCT tarih FROM yanitlar WHERE anket_id=? ORDER BY tarih DESC",(anket_id,)).fetchall()
    liste=[]
    for y in yanitlar:
        v=json.loads(y["veriler"])
        if filtre_arama:
            if not any(filtre_arama in str(val).lower() for val in v.values()):
                continue
        liste.append({"id":y["id"],"tarih":y["tarih"],"saat":y["saat"],"veriler":v})
    # İstatistikler için tüm veriler
    with db() as c:
        tum=c.execute("SELECT veriler FROM yanitlar WHERE anket_id=?",(anket_id,)).fetchall()
    # Soru bazlı istatistik
    soru_istat={}
    for y in tum:
        for k,val in json.loads(y["veriler"]).items():
            if not k.startswith("s_"): continue
            soru_istat.setdefault(k,{})
            for v2 in val.split(","):
                v2=v2.strip()
                if v2: soru_istat[k][v2]=soru_istat[k].get(v2,0)+1
    ctx=gctx()
    ctx.update({"anket":a,"yanitlar":liste,"soru_istat":json.dumps(soru_istat,ensure_ascii=False),
                "tarihler":[r["tarih"] for r in tarihler],
                "filtre_tarih":filtre_tarih,"filtre_arama":filtre_arama,
                "toplam_yanit":len(liste)})
    return render_template("admin_sonuclar.html",**ctx)

@app.route("/admin/yanit_sil/<int:yid>",methods=["POST"])
@giris_gerekli
def yanit_sil(yid):
    with db() as c:
        c.execute("DELETE FROM yanitlar WHERE id=?",(yid,)); c.commit()
    return redirect(request.referrer or url_for("admin_panel"))

# ─── QR Kod ──────────────────────────────────────────────────────
@app.route("/admin/qr/<int:anket_id>")
@giris_gerekli
def qr_kod(anket_id):
    try:
        base_url=request.host_url.rstrip("/")
        url=f"{base_url}/anket/{anket_id}"
        qr=qrcode.QRCode(version=1,box_size=10,border=4)
        qr.add_data(url); qr.make(fit=True)
        img=qr.make_image(fill_color="black",back_color="white")
        buf=io.BytesIO(); img.save(buf,"PNG"); buf.seek(0)
        return send_file(buf,mimetype="image/png",
                         download_name=f"anket_{anket_id}_qr.png",as_attachment=True)
    except ImportError:
        return "qrcode kütüphanesi kurulu değil. pip install qrcode[pil]",500

# ─── Anket CRUD ───────────────────────────────────────────────────
@app.route("/admin/anketler")
@giris_gerekli
def admin_anketler():
    with db() as c:
        anketler=c.execute("SELECT * FROM anketler ORDER BY sira").fetchall()
        sayilar={r["anket_id"]:r["sayi"] for r in
                 c.execute("SELECT anket_id,COUNT(*) as sayi FROM yanitlar GROUP BY anket_id").fetchall()}
    ctx=gctx(); ctx["anketler"]=[dict(a) for a in anketler]; ctx["sayilar"]=sayilar
    return render_template("admin_anketler.html",**ctx)

@app.route("/admin/anket/yeni",methods=["POST"])
@giris_gerekli
def anket_yeni():
    with db() as c:
        aid=c.execute("INSERT INTO anketler (rol,baslik,icon,aktif,sira,aciklama) VALUES (?,?,?,1,99,?)",
                      (request.form.get("rol","diger"),request.form.get("baslik","Yeni Anket"),
                       request.form.get("icon","📝"),request.form.get("aciklama",""))).lastrowid
        c.execute("INSERT INTO bolumler (anket_id,baslik,sira) VALUES (?,?,0)",(aid,"Genel Sorular"))
        c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

@app.route("/admin/anket/<int:aid>/duzenle")
@giris_gerekli
def admin_anket_duzenle(aid):
    a=get_anket(aid)
    if not a: return redirect(url_for("admin_anketler"))
    # Tüm sorular (koşullu soru seçimi için)
    tum_sorular=[]
    for b in a["bolumler"]:
        for s in b["sorular"]:
            if s["tip"] in ["secim","coklu"]:
                tum_sorular.append(s)
    ctx=gctx(); ctx["anket"]=a; ctx["tum_sorular"]=tum_sorular
    return render_template("admin_anket_duzenle.html",**ctx)

@app.route("/admin/anket/<int:aid>/guncelle",methods=["POST"])
@giris_gerekli
def anket_guncelle(aid):
    with db() as c:
        c.execute("""UPDATE anketler SET baslik=?,icon=?,rol=?,aktif=?,
                     aciklama=?,baslangic_tarihi=?,bitis_tarihi=?,gorunum=? WHERE id=?""",
                  (request.form.get("baslik"),request.form.get("icon"),
                   request.form.get("rol"),1 if request.form.get("aktif") else 0,
                   request.form.get("aciklama",""),
                   request.form.get("baslangic_tarihi") or None,
                   request.form.get("bitis_tarihi") or None,
                   request.form.get("gorunum","varsayilan"),
                   aid)); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

@app.route("/admin/anket/<int:aid>/sil",methods=["POST"])
@giris_gerekli
def anket_sil(aid):
    with db() as c:
        c.execute("DELETE FROM yanitlar WHERE anket_id=?",(aid,))
        for b in c.execute("SELECT id FROM bolumler WHERE anket_id=?",(aid,)).fetchall():
            c.execute("DELETE FROM sorular WHERE bolum_id=?",(b["id"],))
        c.execute("DELETE FROM bolumler WHERE anket_id=?",(aid,))
        c.execute("DELETE FROM anketler WHERE id=?",(aid,)); c.commit()
    return redirect(url_for("admin_anketler"))

# ─── Bölüm CRUD ──────────────────────────────────────────────────
@app.route("/admin/bolum/ekle",methods=["POST"])
@giris_gerekli
def bolum_ekle():
    aid=int(request.form["anket_id"])
    with db() as c:
        c.execute("INSERT INTO bolumler (anket_id,baslik,sira) VALUES (?,?,99)",
                  (aid,request.form.get("baslik","Yeni Bölüm"))); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

@app.route("/admin/bolum/<int:bid>/guncelle",methods=["POST"])
@giris_gerekli
def bolum_guncelle(bid):
    aid=int(request.form["anket_id"])
    with db() as c:
        c.execute("UPDATE bolumler SET baslik=? WHERE id=?",(request.form.get("baslik"),bid)); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

@app.route("/admin/bolum/<int:bid>/sil",methods=["POST"])
@giris_gerekli
def bolum_sil(bid):
    aid=int(request.form["anket_id"])
    with db() as c:
        c.execute("DELETE FROM sorular WHERE bolum_id=?",(bid,))
        c.execute("DELETE FROM bolumler WHERE id=?",(bid,)); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

# ─── Soru CRUD ───────────────────────────────────────────────────
@app.route("/admin/soru/ekle",methods=["POST"])
@giris_gerekli
def soru_ekle():
    bid=int(request.form["bolum_id"]); aid=int(request.form["anket_id"])
    tip=request.form.get("tip","yildiz")
    secs=[s.strip() for s in request.form.getlist("secenek") if s.strip()]
    kosul_sid=request.form.get("kosul_soru_id") or None
    kosul_deg=request.form.get("kosul_deger") or None
    if kosul_sid: kosul_sid=int(kosul_sid)
    with db() as c:
        c.execute("""INSERT INTO sorular
                  (bolum_id,metin,tip,secenekler,zorunlu,sira,kosul_soru_id,kosul_deger)
                  VALUES (?,?,?,?,?,99,?,?)""",
                  (bid,request.form.get("metin","Yeni soru?"),tip,
                   json.dumps(secs,ensure_ascii=False),
                   1 if request.form.get("zorunlu") else 0,
                   kosul_sid,kosul_deg)); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

@app.route("/admin/soru/<int:sid>/guncelle",methods=["POST"])
@giris_gerekli
def soru_guncelle(sid):
    aid=int(request.form["anket_id"]); tip=request.form.get("tip","yildiz")
    secs=[s.strip() for s in request.form.getlist("secenek") if s.strip()]
    kosul_sid=request.form.get("kosul_soru_id") or None
    kosul_deg=request.form.get("kosul_deger") or None
    if kosul_sid: kosul_sid=int(kosul_sid)
    with db() as c:
        c.execute("""UPDATE sorular SET metin=?,tip=?,secenekler=?,zorunlu=?,
                     kosul_soru_id=?,kosul_deger=? WHERE id=?""",
                  (request.form.get("metin"),tip,json.dumps(secs,ensure_ascii=False),
                   1 if request.form.get("zorunlu") else 0,
                   kosul_sid,kosul_deg,sid)); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

@app.route("/admin/soru/<int:sid>/sil",methods=["POST"])
@giris_gerekli
def soru_sil(sid):
    aid=int(request.form["anket_id"])
    with db() as c:
        c.execute("DELETE FROM sorular WHERE id=?",(sid,)); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

# ─── Ayarlar ─────────────────────────────────────────────────────
@app.route("/admin/ayarlar",methods=["GET","POST"])
@giris_gerekli
def admin_ayarlar():
    mesaj=None
    if request.method=="POST":
        for k in ["okul_adi","okul_sehir","admin_sifre","tema","anasayfa_gorunum",
                  "hosgeldin_metin","alt_yazi","smtp_host","smtp_port",
                  "smtp_user","smtp_pass","bildirim_email",
                  "anket_varsayilan_gorunum"]:
            v=request.form.get(k)
            if v is not None: ayar_set(k,v)
        ayar_set("bildirim_aktif","1" if request.form.get("bildirim_aktif") else "0")
        ayar_set("anket_gorunum_secim_goster","1" if request.form.get("anket_gorunum_secim_goster") else "0")
        f=request.files.get("amblem")
        if f and f.filename:
            data=f.read(); ext=f.filename.rsplit(".",1)[-1].lower()
            mime="image/png" if ext=="png" else "image/jpeg"
            ayar_set("amblem",f"data:{mime};base64,{base64.b64encode(data).decode()}")
        elif request.form.get("amblem_sil"):
            ayar_set("amblem","")
        mesaj="Ayarlar kaydedildi! ✓"
    ctx=gctx()
    ctx.update({"mesaj":mesaj,"gorunum":ayar("anasayfa_gorunum","kartlar"),
                "anket_varsayilan_gorunum":ayar("anket_varsayilan_gorunum","adim"),
                "anket_gorunum_secim_goster":ayar("anket_gorunum_secim_goster","1"),
                "hosgeldin":ayar("hosgeldin_metin"),"alt_yazi":ayar("alt_yazi"),
                "admin_sifre":ayar("admin_sifre","okul2024"),
                "smtp_host":ayar("smtp_host"),"smtp_port":ayar("smtp_port","587"),
                "smtp_user":ayar("smtp_user"),"smtp_pass":ayar("smtp_pass"),
                "bildirim_email":ayar("bildirim_email"),
                "bildirim_aktif":ayar("bildirim_aktif","0")})
    return render_template("admin_ayarlar.html",**ctx)

# ─── Excel ───────────────────────────────────────────────────────
@app.route("/admin/excel/<int:anket_id>")
@giris_gerekli
def excel_indir(anket_id):
    try:
        import openpyxl
        from openpyxl.styles import Font,PatternFill,Alignment
    except: return "openpyxl kurulu değil.",500
    a=get_anket(anket_id)
    if not a: return redirect(url_for("admin_panel"))
    with db() as c:
        yanitlar=c.execute("SELECT * FROM yanitlar WHERE anket_id=? ORDER BY id",(anket_id,)).fetchall()
    wb=openpyxl.Workbook(); ws=wb.active; ws.title=a["baslik"][:30]
    hdrs=["No","Tarih","Saat"]; sid_order=[]
    for b in a["bolumler"]:
        for s in b["sorular"]:
            hdrs.append(s["metin"][:50]); sid_order.append(s["id"])
    for ci,h in enumerate(hdrs,1):
        cell=ws.cell(1,ci,h); cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor="185FA5")
        cell.alignment=Alignment(horizontal="center",wrap_text=True)
    ws.row_dimensions[1].height=40
    for ri,y in enumerate(yanitlar,2):
        v=json.loads(y["veriler"])
        ws.cell(ri,1,y["id"]); ws.cell(ri,2,y["tarih"]); ws.cell(ri,3,y["saat"])
        for ci,sid in enumerate(sid_order,4):
            ws.cell(ri,ci,v.get(f"s_{sid}",""))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=22
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fn=f"{a['baslik'][:20]}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf,as_attachment=True,download_name=fn,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__=="__main__":
    port=int(os.environ.get("PORT",5000))
    app.run(debug=False,host="0.0.0.0",port=port)
